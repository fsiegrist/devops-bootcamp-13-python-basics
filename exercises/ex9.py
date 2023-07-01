import openpyxl

employees_file = openpyxl.load_workbook("employees.xlsx")
sheet1 = employees_file["Sheet1"]

# remove columns "Job Title" and "Date of Birth"
sheet1.delete_cols(3, 4)

# collect rows
new_rows = []
for employee_row in range(2, sheet1.max_row + 1):
    name = sheet1.cell(employee_row, 1).value
    years_of_experience = sheet1.cell(employee_row, 2).value
    new_rows.append({ 'name': name, 'years_of_experience': int(years_of_experience) })

# sort the rows by "Years of Experience"
new_rows.sort(key=lambda dict: dict.get('years_of_experience'), reverse=True)

# write the sorted rows
for employee_row in range(2, sheet1.max_row + 1):
    index = employee_row - 2
    sheet1.cell(employee_row, 1).value = new_rows[index].get('name')
    sheet1.cell(employee_row, 2).value = new_rows[index].get('years_of_experience')

employees_file.save("employees_sorted.xlsx")
