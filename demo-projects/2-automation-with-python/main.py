import openpyxl
from openpyxl.styles import Font, PatternFill
from copy import copy

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}     # result: { 'AAA Company': 43, 'BBB Company': 17, 'CCC Company': 14 }
total_value_per_supplier = {}  # result: {'AAA Company': 10969059.95, 'BBB Company': 2375499.47, 'CCC Company': 8114363.62}
products_under_10_inv = {}     # result: {25: 7, 30: 6, 74: 2}

for product_row in range(2, product_list.max_row + 1):   # range(75) creates a list of numbers from 0 to 74
    product_num = product_list.cell(product_row, 1).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    supplier_name = product_list.cell(product_row, 4).value

    inventory_price = product_list.cell(product_row, 5)

    # calculate the number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculate the total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # collect products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)

    # calculate and set the value for total inventory price
    inventory_price.value = inventory * price

# set the title of the new column
total_inv_price_title = product_list.cell(1, 5)
total_inv_price_title.value = 'Total Price'
total_inv_price_title.font = copy(product_list.cell(1, 4).font)
total_inv_price_title.fill = copy(product_list.cell(1, 4).fill)

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

inv_file.save("inventory_with_total_value.xlsx")